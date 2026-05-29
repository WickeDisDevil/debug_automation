"""Excel exporter for the Phase-1 categorization showcase.

Layout (per spec):
  * One sheet PER CATEGORY. Each sheet has the five spec columns:
        Issue ID | Title | Category | Priority | Status
  * A leading "Summary" sheet listing every category with its issue count
    so a reviewer can navigate the workbook without scrolling tabs.
  * Sheet names are sanitized for Excel:
      - max 31 chars
      - none of  : \\ / ? * [ ]
      - leading/trailing apostrophes stripped
      - duplicates after sanitization get a numeric suffix

Why one sheet per category (vs. a single grouped sheet):
  Reviewers consume the report category-by-category. Tabs are easier to
  share with a stakeholder ("look at the audio/codec tab") than filtering
  a giant sheet. Excel happily handles hundreds of tabs.

Public surface:
    write_workbook(rows, out_path) -> Path
        Group rows by category and persist as .xlsx. Returns the path
        actually written (the exporter ensures parent dirs exist).
"""

from __future__ import annotations

import re
from collections import OrderedDict, defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bugfix_ai.categorization.schema import CategorizedIssue
from bugfix_ai.config.logging_config import get_logger

log = get_logger(__name__)


# Spec columns, in display order.
SHOWCASE_HEADERS: tuple[str, ...] = (
    "Issue ID",
    "Title",
    "Category",
    "Priority",
    "Status",
)


# ── Sheet-name sanitization ────────────────────────────────────────────────


_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
_SHEET_NAME_MAX = 31


def _sanitize_sheet_name(raw: str) -> str:
    """Make `raw` safe to use as an Excel sheet name."""
    name = _INVALID_SHEET_CHARS.sub("_", raw or "uncategorized").strip()
    name = name.strip("'")
    if not name:
        name = "uncategorized"
    if len(name) > _SHEET_NAME_MAX:
        name = name[:_SHEET_NAME_MAX]
    return name


def _unique_sheet_name(used: set[str], desired: str) -> str:
    """Return a sheet name not already in `used`, suffixing with _2, _3, …."""
    candidate = desired
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = (desired[: _SHEET_NAME_MAX - len(suffix)]) + suffix
        n += 1
    used.add(candidate)
    return candidate


# ── Styling helpers ────────────────────────────────────────────────────────


def _style_header(ws: Worksheet, n_cols: int) -> None:
    fill = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    font = Font(bold=True, color="FFFFFFFF")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, headers: list[str]) -> None:
    """Approximate auto-fit; openpyxl has no true autosize."""
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 80)


# ── Public API ─────────────────────────────────────────────────────────────


def _group_by_category(
    rows: Iterable[CategorizedIssue],
) -> "OrderedDict[str, list[CategorizedIssue]]":
    """Group rows by their `category`. Order: alphabetical, but
    `uncategorized` always sorted last so reviewers see real buckets first."""
    buckets: dict[str, list[CategorizedIssue]] = defaultdict(list)
    for r in rows:
        cat = (r.category or "uncategorized").strip() or "uncategorized"
        buckets[cat].append(r)

    def _sort_key(c: str) -> tuple[int, str]:
        return (1 if c == "uncategorized" else 0, c.lower())

    return OrderedDict((k, buckets[k]) for k in sorted(buckets.keys(), key=_sort_key))


def _write_summary_sheet(
    wb: Workbook, grouped: "OrderedDict[str, list[CategorizedIssue]]"
) -> None:
    ws = wb.active
    assert ws is not None
    ws.title = "Summary"
    headers = ["Category", "Issue Count"]
    ws.append(headers)
    total = 0
    for cat, items in grouped.items():
        ws.append([cat, len(items)])
        total += len(items)
    ws.append(["TOTAL", total])
    _style_header(ws, len(headers))
    _autosize(ws, headers)


def _write_category_sheet(
    wb: Workbook, sheet_name: str, items: list[CategorizedIssue]
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(list(SHOWCASE_HEADERS))
    for r in items:
        sr = r.showcase_row()
        ws.append([sr[h] for h in SHOWCASE_HEADERS])
    _style_header(ws, len(SHOWCASE_HEADERS))
    _autosize(ws, list(SHOWCASE_HEADERS))


def write_workbook(
    rows: Iterable[CategorizedIssue],
    out_path: str | Path,
) -> Path:
    """Write the categorized rows as a multi-sheet .xlsx workbook.

    Parameters
    ----------
    rows:
        Iterable of CategorizedIssue. Order within each category is
        preserved as the input order (caller decides sort order).
    out_path:
        Destination path. Parent directory is created if missing.

    Returns
    -------
    Path
        The resolved output path that was written.
    """
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    grouped = _group_by_category(rows)

    wb = Workbook()
    _write_summary_sheet(wb, grouped)

    used_names: set[str] = {"Summary"}
    for cat, items in grouped.items():
        sheet_name = _unique_sheet_name(used_names, _sanitize_sheet_name(cat))
        _write_category_sheet(wb, sheet_name, items)

    wb.save(out)
    log.info(
        "excel_exporter.wrote",
        path=str(out),
        categories=len(grouped),
        rows=sum(len(v) for v in grouped.values()),
    )
    return out
